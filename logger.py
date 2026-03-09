"""Result logging module - saves to CSV and Excel"""

import os
import csv
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import CSV_FILENAME, EXCEL_FILENAME, FOOD_DISPLAY


class ResultLogger:
    """Logs game results to CSV and Excel files."""

    def __init__(self, output_dir="."):
        self.output_dir = output_dir
        self.csv_path = os.path.join(output_dir, CSV_FILENAME)
        self.excel_path = os.path.join(output_dir, EXCEL_FILENAME)
        self.json_path = os.path.join(output_dir, "results_history.json")
        self.results = []  # In-memory list of all results
        self._load_existing()

    def _load_existing(self):
        """Load existing results from JSON file."""
        if os.path.exists(self.json_path):
            try:
                with open(self.json_path, 'r') as f:
                    self.results = json.load(f)
            except (json.JSONDecodeError, IOError):
                self.results = []

    def add_result(self, food_name, round_number=None, confidence=0.0):
        """Add a new result entry."""
        now = datetime.now()
        if round_number is None:
            round_number = len(self.results) + 1

        entry = {
            "round": round_number,
            "result": food_name,
            "time": now.strftime("%Y-%m-%d %H:%M:%S"),
            "date": now.strftime("%Y-%m-%d"),
            "confidence": round(confidence, 3),
        }
        self.results.append(entry)
        self._append_csv(entry)
        self._save_json()
        return entry

    def _append_csv(self, entry):
        """Append a single entry to the CSV file."""
        file_exists = os.path.exists(self.csv_path)
        with open(self.csv_path, 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=["round", "result", "time", "date", "confidence"])
            if not file_exists:
                writer.writeheader()
            writer.writerow(entry)

    def _save_json(self):
        """Save all results to JSON for persistence."""
        with open(self.json_path, 'w') as f:
            json.dump(self.results, f, indent=2)

    def save_excel(self):
        """Save all results to a formatted Excel file."""
        wb = Workbook()

        # --- Sheet 1: All Results ---
        ws = wb.active
        ws.title = "Results"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        headers = ["Round", "Result", "Time", "Date", "Confidence"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for row_idx, entry in enumerate(self.results, 2):
            ws.cell(row=row_idx, column=1, value=entry["round"]).border = border
            ws.cell(row=row_idx, column=2, value=entry["result"]).border = border
            ws.cell(row=row_idx, column=3, value=entry["time"]).border = border
            ws.cell(row=row_idx, column=4, value=entry["date"]).border = border
            ws.cell(row=row_idx, column=5, value=entry.get("confidence", 0)).border = border

            # Color the result cell
            food = entry["result"]
            if food in FOOD_DISPLAY:
                color = FOOD_DISPLAY[food]["color"].replace("#", "")
                ws.cell(row=row_idx, column=2).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                ws.cell(row=row_idx, column=2).font = Font(bold=True, color="FFFFFF")

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12

        # --- Sheet 2: Statistics ---
        ws2 = wb.create_sheet("Statistics")
        stats = self.get_statistics()

        ws2.cell(row=1, column=1, value="Food Item").font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2, value="Count").font = header_font
        ws2.cell(row=1, column=2).fill = header_fill
        ws2.cell(row=1, column=3, value="Percentage").font = header_font
        ws2.cell(row=1, column=3).fill = header_fill
        ws2.cell(row=1, column=4, value="Last Seen").font = header_font
        ws2.cell(row=1, column=4).fill = header_fill

        for row_idx, (food, data) in enumerate(
            sorted(stats["counts"].items(), key=lambda x: x[1], reverse=True), 2
        ):
            pct = (data / stats["total"] * 100) if stats["total"] > 0 else 0
            ws2.cell(row=row_idx, column=1, value=food.capitalize())
            ws2.cell(row=row_idx, column=2, value=data)
            ws2.cell(row=row_idx, column=3, value=f"{pct:.1f}%")

            # Find last seen
            for entry in reversed(self.results):
                if entry["result"] == food:
                    ws2.cell(row=row_idx, column=4, value=entry["time"])
                    break

        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 10
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 22

        wb.save(self.excel_path)

    def get_statistics(self):
        """Get statistics for all results."""
        total = len(self.results)
        counts = {}
        for entry in self.results:
            food = entry["result"]
            counts[food] = counts.get(food, 0) + 1

        # Hot items (most frequent in last 20 rounds)
        recent = self.results[-20:] if len(self.results) >= 20 else self.results
        recent_counts = {}
        for entry in recent:
            food = entry["result"]
            recent_counts[food] = recent_counts.get(food, 0) + 1

        # Streaks
        streaks = {}
        if self.results:
            current_food = self.results[-1]["result"]
            streak = 0
            for entry in reversed(self.results):
                if entry["result"] == current_food:
                    streak += 1
                else:
                    break
            streaks["current"] = {"food": current_food, "count": streak}

        # Missing items (items not seen in last N rounds)
        last_n = 50
        recent_foods = set(e["result"] for e in self.results[-last_n:])
        from config import FOOD_ITEMS
        cold_items = [f for f in FOOD_ITEMS if f not in recent_foods]

        return {
            "total": total,
            "counts": counts,
            "recent_counts": recent_counts,
            "recent_results": [e["result"] for e in self.results[-30:]],
            "streaks": streaks,
            "cold_items": cold_items,
            "last_result": self.results[-1]["result"] if self.results else None,
        }

    def get_last_n_results(self, n=30):
        """Get the last N results."""
        return self.results[-n:]

    @property
    def total_rounds(self):
        return len(self.results)
